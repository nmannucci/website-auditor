#!/usr/bin/env python3
"""
Batch Website Auditor
Processes multiple CPA/accountant websites from a CSV or Excel file
Supports parallel processing for faster batch audits
"""

import csv
import os
import sys
import argparse
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Dict
from concurrent.futures import ThreadPoolExecutor, as_completed
from website_auditor import WebsiteAuditor

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class BatchAuditor:
    def __init__(self, max_workers: int = 3):
        self.max_workers = max_workers
        self.results_dir = Path("batch_results")
        self.results_dir.mkdir(exist_ok=True)
        # Thread-safe progress tracking
        self._progress_lock = threading.Lock()
        self._completed_count = 0
        self._total_count = 0

    def _audit_single_site(self, site_info: Dict, idx: int) -> Dict:
        """Worker function to audit a single site (thread-safe)"""
        # Each thread gets its own WebsiteAuditor instance
        auditor = WebsiteAuditor()

        url = site_info['url']
        company_name = site_info['company_name'] or url

        try:
            audit_result = auditor.audit_website(url, company_name=company_name)
            audit_result['input_notes'] = site_info['notes']

            # Update progress
            with self._progress_lock:
                self._completed_count += 1
                completed = self._completed_count
                total = self._total_count

            if 'error' not in audit_result:
                rec = audit_result['recommendation']
                print(f"\n✅ [{completed}/{total}] {company_name}: {rec['recommendation']} - Score: {rec['score']}/105")
            else:
                print(f"\n⚠️  [{completed}/{total}] {company_name}: Error - {audit_result['error'][:50]}")

            return audit_result

        except Exception as e:
            with self._progress_lock:
                self._completed_count += 1
                completed = self._completed_count
                total = self._total_count

            print(f"\n❌ [{completed}/{total}] {company_name}: Failed - {str(e)[:50]}")
            return {
                'url': url,
                'company_name': company_name,
                'error': str(e)
            }

    def _read_csv(self, file_path: str) -> List[Dict]:
        """Read URLs from a CSV file"""
        urls = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if 'url' in row and row['url'].strip():
                    urls.append({
                        'url': row['url'].strip(),
                        'company_name': row.get('company_name', '').strip(),
                        'notes': row.get('notes', '').strip()
                    })
        return urls

    def _read_excel(self, file_path: str) -> List[Dict]:
        """Read URLs from an Excel file (.xlsx, .xls)"""
        if not EXCEL_SUPPORT:
            print("❌ Error: Excel support not installed. Run: pip install openpyxl")
            return []

        urls = []
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # Get header row to find column indices
        headers = {}
        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
            if cell:
                headers[str(cell).lower().strip()] = col_idx

        if 'url' not in headers:
            print("❌ Error: Excel file must have a 'url' column")
            wb.close()
            return []

        url_idx = headers['url']
        company_idx = headers.get('company_name')
        notes_idx = headers.get('notes')

        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            url_value = row[url_idx] if url_idx < len(row) else None
            if url_value and str(url_value).strip():
                urls.append({
                    'url': str(url_value).strip(),
                    'company_name': str(row[company_idx]).strip() if company_idx is not None and company_idx < len(row) and row[company_idx] else '',
                    'notes': str(row[notes_idx]).strip() if notes_idx is not None and notes_idx < len(row) and row[notes_idx] else ''
                })

        wb.close()
        return urls

    def _update_original_csv(self, file_path: str, results: List[Dict]) -> None:
        """Update original CSV file with score and recommendation columns"""
        # Build lookup by URL
        results_by_url = {}
        for r in results:
            url = r.get('url', '').strip().lower().rstrip('/')
            results_by_url[url] = r

        # Read original file
        rows = []
        fieldnames = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            fieldnames = list(reader.fieldnames) if reader.fieldnames else []
            rows = list(reader)

        # Add new columns if not present
        if 'audit_score' not in fieldnames:
            fieldnames.append('audit_score')
        if 'audit_recommendation' not in fieldnames:
            fieldnames.append('audit_recommendation')

        # Update rows with results
        for row in rows:
            url = row.get('url', '').strip().lower().rstrip('/')
            if url in results_by_url:
                result = results_by_url[url]
                if 'recommendation' in result:
                    row['audit_score'] = result['recommendation'].get('score', '')
                    row['audit_recommendation'] = result['recommendation'].get('recommendation', '')
                else:
                    row['audit_score'] = 'ERROR'
                    row['audit_recommendation'] = result.get('error', 'Unknown error')[:50]

        # Write back
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def _update_original_excel(self, file_path: str, results: List[Dict]) -> None:
        """Update original Excel file with score and recommendation columns"""
        if not EXCEL_SUPPORT:
            print("❌ Error: Excel support not installed. Run: pip install openpyxl")
            return

        # Build lookup by URL
        results_by_url = {}
        for r in results:
            url = r.get('url', '').strip().lower().rstrip('/')
            results_by_url[url] = r

        # Open workbook for editing (not read_only)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Get headers and find url column
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).lower().strip()] = col_idx

        url_col = headers.get('url')
        if not url_col:
            print("❌ Error: Could not find 'url' column in Excel file")
            wb.close()
            return

        # Find or create audit_score and audit_recommendation columns
        max_col = ws.max_column
        score_col = headers.get('audit_score')
        rec_col = headers.get('audit_recommendation')

        if not score_col:
            score_col = max_col + 1
            ws.cell(row=1, column=score_col, value='audit_score')
            max_col += 1

        if not rec_col:
            rec_col = max_col + 1
            ws.cell(row=1, column=rec_col, value='audit_recommendation')

        # Update rows with results
        for row_idx in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=row_idx, column=url_col).value
            if url_cell:
                url = str(url_cell).strip().lower().rstrip('/')
                if url in results_by_url:
                    result = results_by_url[url]
                    if 'recommendation' in result:
                        ws.cell(row=row_idx, column=score_col, value=result['recommendation'].get('score', ''))
                        ws.cell(row=row_idx, column=rec_col, value=result['recommendation'].get('recommendation', ''))
                    else:
                        ws.cell(row=row_idx, column=score_col, value='ERROR')
                        ws.cell(row=row_idx, column=rec_col, value=str(result.get('error', 'Unknown'))[:50])

        # Save
        wb.save(file_path)
        wb.close()

    def _update_original_file(self, file_path: str, results: List[Dict]) -> None:
        """Update the original input file with audit results"""
        file_ext = Path(file_path).suffix.lower()

        if file_ext == '.csv':
            self._update_original_csv(file_path, results)
        elif file_ext in ['.xlsx', '.xls']:
            self._update_original_excel(file_path, results)

        print(f"📝 Updated original file with audit results: {file_path}")

    def process_file(self, file_path: str, parallel: bool = True) -> List[Dict]:
        """Process all URLs from a CSV or Excel file

        Args:
            file_path: Path to CSV or Excel file with URLs
            parallel: If True, process sites in parallel (faster). Default True.
        """

        if not os.path.exists(file_path):
            print(f"❌ Error: File not found: {file_path}")
            return []

        # Determine file type and read URLs
        file_ext = Path(file_path).suffix.lower()
        if file_ext in ['.xlsx', '.xls']:
            urls = self._read_excel(file_path)
        elif file_ext == '.csv':
            urls = self._read_csv(file_path)
        else:
            print(f"❌ Error: Unsupported file type: {file_ext}")
            print("Supported formats: .csv, .xlsx, .xls")
            return []

        if not urls:
            print("❌ No URLs found in file")
            print("Make sure your file has a 'url' column")
            return []

        # Reset progress tracking
        self._completed_count = 0
        self._total_count = len(urls)

        print(f"\n{'='*70}")
        print(f"🚀 BATCH AUDIT STARTED")
        print(f"{'='*70}")
        print(f"Total websites to audit: {len(urls)}")
        print(f"Processing mode: {'Parallel (' + str(self.max_workers) + ' workers)' if parallel else 'Sequential'}")
        print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*70}")

        start_time = datetime.now()

        if parallel and len(urls) > 1:
            # Parallel processing
            results = self._process_parallel(urls)
        else:
            # Sequential processing (for single URL or if parallel disabled)
            results = self._process_sequential(urls)

        elapsed = (datetime.now() - start_time).total_seconds()

        # Update original file with results
        self._update_original_file(file_path, results)

        # Generate summary report
        summary_path = self._generate_summary_report(results, file_path)

        print(f"\n{'='*70}")
        print(f"✅ BATCH AUDIT COMPLETE")
        print(f"{'='*70}")
        print(f"Total audited: {len(results)}")
        print(f"Time elapsed: {elapsed:.1f}s ({elapsed/len(results):.1f}s per site)")
        print(f"Summary saved to: {summary_path}")
        print(f"{'='*70}\n")

        return results

    def _process_parallel(self, urls: List[Dict]) -> List[Dict]:
        """Process URLs in parallel using ThreadPoolExecutor"""
        results = []

        print(f"\n🔄 Starting {self.max_workers} parallel workers...")

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks
            future_to_site = {
                executor.submit(self._audit_single_site, site_info, idx): site_info
                for idx, site_info in enumerate(urls, 1)
            }

            # Collect results as they complete
            for future in as_completed(future_to_site):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    site_info = future_to_site[future]
                    print(f"\n❌ Unexpected error for {site_info['url']}: {str(e)}")
                    results.append({
                        'url': site_info['url'],
                        'company_name': site_info.get('company_name', ''),
                        'error': str(e)
                    })

        return results

    def _process_sequential(self, urls: List[Dict]) -> List[Dict]:
        """Process URLs sequentially (original behavior)"""
        results = []

        for idx, site_info in enumerate(urls, 1):
            result = self._audit_single_site(site_info, idx)
            results.append(result)

        return results

    def _generate_summary_report(self, results: List[Dict], input_file: str) -> Path:
        """Generate CSV and Markdown summary of batch results"""

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # CSV Summary
        csv_summary_path = self.results_dir / f"summary_{timestamp}.csv"

        with open(csv_summary_path, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                'company_name',
                'url',
                'recommendation',
                'score',
                'percentage',
                'total_issues',
                'has_clear_cta',
                'has_contact_form',
                'has_phone_number',
                'has_team_info',
                'has_credentials',
                'has_google_maps',
                'design_score',
                'load_time_seconds',
                'has_valid_ssl',
                'ssl_expiry',
                'report_path',
                'pdf_path',
                'error'
            ]

            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for result in results:
                if 'error' in result and 'recommendation' not in result:
                    # Error case
                    writer.writerow({
                        'company_name': result.get('company_name', ''),
                        'url': result.get('url', ''),
                        'error': result.get('error', 'Unknown error')
                    })
                else:
                    # Successful audit
                    rec = result['recommendation']
                    sections = result['audit_sections']

                    # Get SSL info
                    ssl_info = sections['technical'].get('ssl', {})

                    writer.writerow({
                        'company_name': result.get('company_name', ''),
                        'url': result['url'],
                        'recommendation': rec['recommendation'],
                        'score': rec['score'],
                        'percentage': rec['percentage'],
                        'total_issues': rec['total_issues'],
                        'has_clear_cta': sections['conversion_elements']['has_clear_cta'],
                        'has_contact_form': sections['conversion_elements']['has_contact_form'],
                        'has_phone_number': sections['conversion_elements']['has_phone_number'],
                        'has_team_info': sections['trust_signals']['has_team_info'],
                        'has_credentials': sections['trust_signals']['has_credentials'],
                        'has_google_maps': sections['trust_signals']['has_google_maps'],
                        'design_score': sections['visual_design']['score'],
                        'load_time_seconds': sections['technical']['load_time_seconds'],
                        'has_valid_ssl': ssl_info.get('is_valid', False),
                        'ssl_expiry': ssl_info.get('expiry_date', ''),
                        'report_path': result.get('report_path', ''),
                        'pdf_path': result.get('pdf_path', ''),
                        'error': ''
                    })

        # Markdown Summary
        md_summary_path = self.results_dir / f"summary_{timestamp}.md"

        with open(md_summary_path, 'w', encoding='utf-8') as f:
            f.write(f"# Batch Audit Summary Report\n\n")
            f.write(f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"**Input File:** {input_file}\n")
            f.write(f"**Total Sites Audited:** {len(results)}\n\n")

            # Categorize results
            strong_yes = [r for r in results if r.get('recommendation', {}).get('recommendation') == 'STRONG YES']
            yes = [r for r in results if r.get('recommendation', {}).get('recommendation') == 'YES']
            maybe = [r for r in results if r.get('recommendation', {}).get('recommendation') == 'MAYBE']
            no = [r for r in results if r.get('recommendation', {}).get('recommendation') == 'NO']
            errors = [r for r in results if 'error' in r and 'recommendation' not in r]

            f.write(f"## 📊 Overview\n\n")
            f.write(f"| Category | Count | Percentage |\n")
            f.write(f"|----------|-------|------------|\n")
            f.write(f"| 🔥 STRONG YES | {len(strong_yes)} | {len(strong_yes)/len(results)*100:.1f}% |\n")
            f.write(f"| ✅ YES | {len(yes)} | {len(yes)/len(results)*100:.1f}% |\n")
            f.write(f"| 🤔 MAYBE | {len(maybe)} | {len(maybe)/len(results)*100:.1f}% |\n")
            f.write(f"| ❌ NO | {len(no)} | {len(no)/len(results)*100:.1f}% |\n")
            if errors:
                f.write(f"| ⚠️ ERRORS | {len(errors)} | {len(errors)/len(results)*100:.1f}% |\n")

            # Top Prospects
            if strong_yes or yes:
                f.write(f"\n## 🎯 Top Prospects (STRONG YES & YES)\n\n")

                top_prospects = strong_yes + yes
                # Sort by score (lowest first = most opportunity)
                top_prospects.sort(key=lambda x: x.get('recommendation', {}).get('score', 100))

                for result in top_prospects:
                    rec = result['recommendation']
                    company = result.get('company_name', result['url'])

                    f.write(f"### {company}\n\n")
                    f.write(f"- **URL:** {result['url']}\n")
                    f.write(f"- **Recommendation:** {rec['recommendation']}\n")
                    f.write(f"- **Score:** {rec['score']}/100 ({rec['percentage']}%)\n")
                    f.write(f"- **Issues Found:** {rec['total_issues']}\n")
                    f.write(f"- **Reason:** {rec['reason']}\n")

                    # Top 3 opportunities
                    if rec['opportunities']:
                        f.write(f"- **Top Opportunities:**\n")
                        for opp in rec['opportunities'][:3]:
                            f.write(f"  - {opp}\n")

                    f.write(f"- **Reports:** [Markdown]({result.get('report_path', '')}) | [PDF]({result.get('pdf_path', '')})\n")
                    f.write(f"\n---\n\n")

            # Maybe prospects
            if maybe:
                f.write(f"\n## 🤔 Moderate Prospects (MAYBE)\n\n")
                for result in maybe:
                    rec = result['recommendation']
                    company = result.get('company_name', result['url'])
                    f.write(f"- **{company}** - Score: {rec['score']}/100 - [{result.get('report_path', 'Report')}]({result.get('report_path', '')})\n")

            # Low priority
            if no:
                f.write(f"\n## ❌ Low Priority (NO)\n\n")
                for result in no:
                    rec = result['recommendation']
                    company = result.get('company_name', result['url'])
                    f.write(f"- **{company}** - Score: {rec['score']}/100 - Well optimized\n")

            # Errors
            if errors:
                f.write(f"\n## ⚠️ Failed Audits\n\n")
                for result in errors:
                    company = result.get('company_name', result.get('url', 'Unknown'))
                    f.write(f"- **{company}** - Error: {result.get('error', 'Unknown error')}\n")

            # Export info
            f.write(f"\n## 📁 Files Generated\n\n")
            f.write(f"- **CSV Summary:** {csv_summary_path}\n")
            f.write(f"- **Individual Reports:** See `reports/` directory\n")
            f.write(f"- **Screenshots:** See `screenshots/` directory\n\n")

        print(f"📄 CSV summary: {csv_summary_path}")
        print(f"📄 Markdown summary: {md_summary_path}")

        return md_summary_path


def main():
    """CLI entry point for batch processing"""

    parser = argparse.ArgumentParser(
        description='Batch Website Auditor - Process multiple websites in parallel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python batch_auditor.py prospects.csv                    # Default: 3 parallel workers
  python batch_auditor.py prospects.xlsx --workers 5      # Use 5 parallel workers
  python batch_auditor.py prospects.csv --sequential      # Process one at a time

File format:
  CSV or Excel file with columns: url (required), company_name (optional), notes (optional)
        """
    )

    parser.add_argument('file', help='CSV or Excel file with URLs to audit')
    parser.add_argument(
        '-w', '--workers',
        type=int,
        default=3,
        help='Number of parallel workers (default: 3, max recommended: 5)'
    )
    parser.add_argument(
        '-s', '--sequential',
        action='store_true',
        help='Process sites one at a time (disables parallel processing)'
    )

    args = parser.parse_args()

    # Check for API key
    if not os.getenv("ANTHROPIC_API_KEY"):
        print("❌ Error: ANTHROPIC_API_KEY not found in environment")
        print("Please create a .env file with your API key:")
        print("  ANTHROPIC_API_KEY=your_key_here")
        sys.exit(1)

    # Validate workers
    if args.workers < 1:
        print("❌ Error: --workers must be at least 1")
        sys.exit(1)
    if args.workers > 10:
        print("⚠️  Warning: More than 10 workers may cause rate limiting or system issues")

    batch_auditor = BatchAuditor(max_workers=args.workers)
    results = batch_auditor.process_file(args.file, parallel=not args.sequential)

    # Show final summary
    if results:
        strong_yes = sum(1 for r in results if r.get('recommendation', {}).get('recommendation') == 'STRONG YES')
        yes_count = sum(1 for r in results if r.get('recommendation', {}).get('recommendation') == 'YES')

        print(f"\n🎯 ACTION ITEMS:")
        print(f"   - {strong_yes} STRONG YES prospects (high priority)")
        print(f"   - {yes_count} YES prospects (good priority)")
        print(f"   - Check batch_results/ folder for detailed summaries\n")


if __name__ == "__main__":
    main()
