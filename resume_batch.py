#!/usr/bin/env python3
"""
Resume Batch Audit - Continue auditing from where it left off
Skips rows that already have audit scores
"""

import os
import sys
import openpyxl
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Add the current directory to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from website_auditor import WebsiteAuditor


def get_pending_urls(file_path: str):
    """Get URLs that haven't been audited yet (no audit_score)"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    headers = [str(cell.value) if cell.value else '' for cell in ws[1]]

    # Find column indices
    url_idx = headers.index('url') if 'url' in headers else -1
    name_idx = 6  # company_name
    score_idx = headers.index('audit_score') if 'audit_score' in headers else -1

    pending = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        url = row[url_idx] if url_idx >= 0 and url_idx < len(row) else ''
        name = row[name_idx] if name_idx >= 0 and name_idx < len(row) else ''
        score = row[score_idx] if score_idx >= 0 and score_idx < len(row) else None

        if url and score is None:
            pending.append({
                'row_num': row_num,
                'url': str(url).strip(),
                'company_name': str(name).strip() if name else ''
            })

    wb.close()
    return pending


def update_excel_row(file_path: str, row_num: int, score, recommendation):
    """Update a single row in the Excel file with audit results"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = [str(cell.value) if cell.value else '' for cell in ws[1]]

    # Find or create score and recommendation columns
    score_col = headers.index('audit_score') + 1 if 'audit_score' in headers else None
    rec_col = headers.index('audit_recommendation') + 1 if 'audit_recommendation' in headers else None

    if not score_col:
        score_col = ws.max_column + 1
        ws.cell(row=1, column=score_col, value='audit_score')

    if not rec_col:
        rec_col = ws.max_column + 1
        ws.cell(row=1, column=rec_col, value='audit_recommendation')

    # Update the row
    ws.cell(row=row_num, column=score_col, value=score)
    ws.cell(row=row_num, column=rec_col, value=recommendation)

    wb.save(file_path)
    wb.close()


def audit_site(site_info, progress_lock, progress, total, file_path):
    """Audit a single site and update the Excel file"""
    auditor = WebsiteAuditor()

    url = site_info['url']
    company_name = site_info['company_name'] or url
    row_num = site_info['row_num']

    try:
        result = auditor.audit_website(url, company_name=company_name)

        with progress_lock:
            progress[0] += 1
            current = progress[0]

        if 'error' not in result:
            score = result['recommendation']['score']
            rec = result['recommendation']['recommendation']
            print(f"✅ [{current}/{total}] {company_name}: {rec} - Score: {score}/105")

            # Update Excel file
            update_excel_row(file_path, row_num, score, rec)

            return {'success': True, 'row_num': row_num, 'score': score, 'rec': rec}
        else:
            print(f"⚠️  [{current}/{total}] {company_name}: Error - {result['error'][:50]}")
            update_excel_row(file_path, row_num, 'ERROR', result['error'][:50])
            return {'success': False, 'row_num': row_num, 'error': result['error']}

    except Exception as e:
        with progress_lock:
            progress[0] += 1
            current = progress[0]

        print(f"❌ [{current}/{total}] {company_name}: Failed - {str(e)[:50]}")
        update_excel_row(file_path, row_num, 'ERROR', str(e)[:50])
        return {'success': False, 'row_num': row_num, 'error': str(e)}


def main():
    file_path = "csv-batches/2- Outscraper - Accountants (Pasadena, CA)-Verified.xlsx"
    max_workers = 3

    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        sys.exit(1)

    # Get pending URLs
    pending = get_pending_urls(file_path)

    if not pending:
        print("✅ All URLs have already been audited!")
        sys.exit(0)

    print(f"\n{'='*70}")
    print(f"🔄 RESUMING BATCH AUDIT")
    print(f"{'='*70}")
    print(f"File: {file_path}")
    print(f"Pending audits: {len(pending)}")
    print(f"Workers: {max_workers}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    # Progress tracking
    progress_lock = threading.Lock()
    progress = [0]  # Use list for mutable reference
    total = len(pending)

    start_time = datetime.now()

    # Process in parallel
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(audit_site, site, progress_lock, progress, total, file_path): site
            for site in pending
        }

        for future in as_completed(futures):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                print(f"❌ Unexpected error: {e}")

    elapsed = (datetime.now() - start_time).total_seconds()

    # Summary
    successful = sum(1 for r in results if r.get('success'))
    failed = len(results) - successful

    print(f"\n{'='*70}")
    print(f"✅ BATCH AUDIT COMPLETE")
    print(f"{'='*70}")
    print(f"Processed: {len(results)}")
    print(f"Successful: {successful}")
    print(f"Failed: {failed}")
    print(f"Time elapsed: {elapsed:.1f}s ({elapsed/len(results):.1f}s per site)")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
